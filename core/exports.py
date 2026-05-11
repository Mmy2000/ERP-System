import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill('solid', start_color='1E293B')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center')

DATA_FONT = Font(name='Arial', size=10)
DATA_ALIGN = Alignment(vertical='center')

ALT_FILL = PatternFill('solid', start_color='F8FAFC')

THIN_BORDER = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0'),
)


def build_workbook(sheets: list[dict]) -> Workbook:
    """
    sheets = [
        {
            'title': 'Products',
            'headers': ['SKU', 'Name', ...],
            'rows': [[val, val, ...], ...],
            'col_widths': [15, 30, ...],   # optional
            'number_formats': {3: '#,##0.00', 4: '#,##0.00'},  # optional, 1-based col index
        }
    ]
    """
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_def in sheets:
        ws = wb.create_sheet(title=sheet_def['title'])
        headers = sheet_def['headers']
        rows = sheet_def['rows']
        col_widths = sheet_def.get('col_widths', [])
        num_fmts = sheet_def.get('number_formats', {})

        ws.row_dimensions[1].height = 28

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER

        for row_idx, row_data in enumerate(rows, 2):
            fill = ALT_FILL if row_idx % 2 == 0 else None
            ws.row_dimensions[row_idx].height = 20
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = DATA_FONT
                cell.alignment = DATA_ALIGN
                cell.border = THIN_BORDER
                if fill:
                    cell.fill = fill
                if col_idx in num_fmts:
                    cell.number_format = num_fmts[col_idx]

        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    return wb


def workbook_response(wb: Workbook, filename: str) -> HttpResponse:
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response